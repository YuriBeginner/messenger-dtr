from flask import Flask, request, send_from_directory
from openpyxl import Workbook, load_workbook
from datetime import datetime, timezone, timedelta
import os
import json
import base64
import requests

app = Flask(__name__)

VERIFY_TOKEN = "ojt_dtr_token"

# =========================================================
# ------------------- GITHUB CONFIG -----------------------
# =========================================================

GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN")
GITHUB_REPO = "YuriBeginner/messenger-dtr"
GITHUB_BRANCH = "main"

GITHUB_HEADERS = {
    "Authorization": f"token {GITHUB_TOKEN}",
    "Accept": "application/vnd.github.v3+json"
}

def github_get_json(filename):
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{filename}"
    params = {"ref": GITHUB_BRANCH}

    r = requests.get(url, headers=GITHUB_HEADERS, params=params)

    if r.status_code == 200:
        data = r.json()
        content = base64.b64decode(data["content"]).decode().strip()

        if not content:
            # If file somehow empty
            return {} if filename == "users.json" else [], data["sha"]

        try:
            return json.loads(content), data["sha"]
        except:
            print("Invalid JSON detected in", filename)
            return {} if filename == "users.json" else [], data["sha"]

    elif r.status_code == 404:
        return {} if filename == "users.json" else [], None

    else:
        print("GitHub GET error:", r.text)
        return {} if filename == "users.json" else [], None



def github_save_json(filename, content_data, sha):
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{filename}"

    encoded_content = base64.b64encode(
        json.dumps(content_data, indent=2).encode()
    ).decode()

    payload = {
        "message": f"Update {filename}",
        "content": encoded_content,
        "branch": GITHUB_BRANCH
    }

    if sha:
        payload["sha"] = sha

    r = requests.put(url, headers=GITHUB_HEADERS, json=payload)

    if r.status_code not in [200, 201]:
        print("GitHub SAVE error:", r.text)

# =========================================================
# ------------------- JSON STORAGE ------------------------
# =========================================================

def load_users():
    users, sha = github_get_json("users.json")
    return users, sha

def save_users(users, sha):
    github_save_json("users.json", users, sha)


def load_processed():
    data, sha = github_get_json("processed_messages.json")
    return set(data), sha

def save_processed(processed_set, sha):
    github_save_json("processed_messages.json", list(processed_set), sha)

# =========================================================
# ------------------- EXCEL STORAGE -----------------------
# =========================================================

if not os.path.exists("DTR"):
    os.makedirs("DTR")

def get_file(name):
    return f"DTR/{name}.xlsx"

def download_from_github(name):
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/DTR/{name}.xlsx"

    r = requests.get(url, headers=GITHUB_HEADERS)

    if r.status_code == 200:
        data = r.json()
        content = base64.b64decode(data["content"])

        with open(get_file(name), "wb") as f:
            f.write(content)

def ensure_file(name):
    f = get_file(name)

    if not os.path.exists(f):
        # Try to download existing file first
        download_from_github(name)

    if not os.path.exists(f):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Date", "Time In", "Time Out"])
        wb.save(f)

def is_empty(cell):
    return cell.value is None or str(cell.value).strip() == ""

def upload_to_github(file_path, name):
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/DTR/{name}.xlsx"

    with open(file_path, "rb") as f:
        content = base64.b64encode(f.read()).decode()

    r = requests.get(url, headers=GITHUB_HEADERS)
    sha = r.json().get("sha") if r.status_code == 200 else None

    payload = {
        "message": f"Update DTR for {name}",
        "content": content,
        "branch": GITHUB_BRANCH
    }

    if sha:
        payload["sha"] = sha

    requests.put(url, headers=GITHUB_HEADERS, json=payload)

# =========================================================
# ------------------- MESSENGER ---------------------------
# =========================================================

PAGE_ACCESS_TOKEN = os.environ.get("PAGE_ACCESS_TOKEN")

def send_message(psid, message):
    url = f"https://graph.facebook.com/v19.0/me/messages?access_token={PAGE_ACCESS_TOKEN}"

    payload = {
        "recipient": {"id": psid},
        "message": {"text": message}
    }

    r = requests.post(url, json=payload)
    print("Messenger reply:", r.text)

# =========================================================
# ------------------- TIME LOGIC --------------------------
# =========================================================

def log_time(name, action, timestamp):
    ensure_file(name)
    f = get_file(name)
    wb = load_workbook(f)
    ws = wb.active

    utc_time = datetime.fromtimestamp(timestamp / 1000, tz=timezone.utc)
    ph_time = utc_time.astimezone(timezone(timedelta(hours=8)))

    date_str = ph_time.strftime("%Y-%m-%d")
    time_str = ph_time.strftime("%H:%M:%S")

    for row in ws.iter_rows(min_row=2):
        if row[0].value == name and row[1].value == date_str:

            if action == "TIME IN":
                return False

            if action == "TIME OUT" and is_empty(row[3]):
                row[3].value = time_str
                wb.save(f)
                upload_to_github(f, name)
                return True

            return False

    if action == "TIME IN":
        ws.append([name, date_str, time_str, None])
        wb.save(f)
        upload_to_github(f, name)
        return True

    return False

# =========================================================
# ------------------- WEBHOOK VERIFY ----------------------
# =========================================================

@app.route("/webhook", methods=["GET"])
def verify():
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    if token == VERIFY_TOKEN:
        return challenge
    return "Verification failed"

# =========================================================
# ------------------- WEBHOOK POST ------------------------
# =========================================================

@app.route("/webhook", methods=["POST"])
def webhook():
    data = request.json

    try:
        entry = data["entry"][0]
        messaging = entry["messaging"][0]

        if "message" not in messaging:
            return "ok", 200

        message_id = messaging["message"]["mid"]

        # ----- Deduplication -----
        processed, processed_sha = load_processed()

        if message_id in processed:
            print("Duplicate ignored:", message_id)
            return "ok", 200

        processed.add(message_id)
        save_processed(processed, processed_sha)

        sender_id = messaging["sender"]["id"]
        raw_text = messaging["message"]["text"].strip()
        text = raw_text.upper()
        timestamp = messaging["timestamp"]

        users, users_sha = load_users()

        # ----- REGISTER -----
        if text.startswith("REGISTER "):
            if sender_id in users:
                send_message(sender_id, "⚠️ You are already registered.")
                return "ok", 200

            real_name = raw_text.replace("REGISTER ", "").strip()
            users[sender_id] = real_name
            save_users(users, users_sha)

            send_message(sender_id, f"✅ Successfully registered as {real_name}")
            return "ok", 200

        # ----- FIXNAME -----
        if text.startswith("FIXNAME "):
            new_name = raw_text[8:].strip()
            old_name = users.get(sender_id)

            if old_name:
                old_file = get_file(old_name)
                new_file = get_file(new_name)

                if os.path.exists(old_file):
                    os.rename(old_file, new_file)
                    upload_to_github(new_file, new_name)

                users[sender_id] = new_name
                save_users(users, users_sha)

                send_message(sender_id, f"✅ Name updated to {new_name}")

            return "ok", 200

        # ----- TIME IN / OUT -----
        if text in ["TIME IN", "TIME OUT"]:

            name = users.get(sender_id)

            if not name:
                send_message(sender_id, "⚠️ Please REGISTER first:\nREGISTER Your Full Name")
                return "ok", 200

            success = log_time(name, text, timestamp)

            utc_time = datetime.fromtimestamp(timestamp / 1000, tz=timezone.utc)
            ph_time = utc_time.astimezone(timezone(timedelta(hours=8)))
            time_str = ph_time.strftime("%H:%M:%S")

            if success:
                send_message(sender_id, f"✅ {text} recorded at {time_str}")
            else:
                send_message(sender_id, f"⚠️ {text} already recorded or invalid.")

            return "ok", 200

    except Exception as e:
        print("Error:", e)

    return "ok", 200

# =========================================================
# ------------------- DOWNLOAD ----------------------------
# =========================================================

@app.route("/download/<name>")
def download_file(name):
    return send_from_directory("DTR", f"{name}.xlsx", as_attachment=True)

# =========================================================
# ------------------- PRIVACY -----------------------------
# =========================================================

@app.route("/privacy.html")
def privacy():
    return send_from_directory('.', 'privacy.html')

@app.route("/")
def home():
    return "OJT DTR Bot is running!"




